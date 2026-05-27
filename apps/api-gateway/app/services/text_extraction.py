import csv
from dataclasses import dataclass
from io import BytesIO, StringIO
from pathlib import Path
from typing import Iterable


@dataclass(frozen=True)
class ExtractedPage:
    page_number: int
    text: str


class TextExtractionError(ValueError):
    pass


def extract_text_pages(filename: str, data: bytes) -> list[ExtractedPage]:
    extension = Path(filename).suffix.lower()
    if extension == ".pdf":
        return _extract_pdf(data)
    if extension == ".txt":
        return [_single_page(_decode_text(data))]
    if extension == ".csv":
        return [_single_page(_extract_csv(data))]
    if extension == ".xlsx":
        return [_single_page(_extract_xlsx(data))]
    raise TextExtractionError(f"Unsupported file extension: {extension}")


def flatten_pages(pages: Iterable[ExtractedPage]) -> str:
    return "\n\n".join(page.text for page in pages if page.text.strip())


def _extract_pdf(data: bytes) -> list[ExtractedPage]:
    try:
        from pypdf import PdfReader

        reader = PdfReader(BytesIO(data))
        pages = [
            ExtractedPage(page_number=index + 1, text=page.extract_text() or "")
            for index, page in enumerate(reader.pages)
        ]
    except Exception as exc:
        raise TextExtractionError("Failed to extract text from PDF.") from exc

    return _ensure_text(pages)


def _extract_csv(data: bytes) -> str:
    text = _decode_text(data)
    rows = csv.reader(StringIO(text))
    return "\n".join(", ".join(cell.strip() for cell in row) for row in rows)


def _extract_xlsx(data: bytes) -> str:
    try:
        from openpyxl import load_workbook

        workbook = load_workbook(BytesIO(data), read_only=True, data_only=True)
        lines: list[str] = []
        for worksheet in workbook.worksheets:
            lines.append(f"Sheet: {worksheet.title}")
            for row in worksheet.iter_rows(values_only=True):
                values = ["" if value is None else str(value).strip() for value in row]
                if any(values):
                    lines.append(", ".join(values))
    except Exception as exc:
        raise TextExtractionError("Failed to extract text from XLSX.") from exc

    return "\n".join(lines)


def _decode_text(data: bytes) -> str:
    try:
        return data.decode("utf-8-sig")
    except UnicodeDecodeError as exc:
        raise TextExtractionError("Uploaded text file must be UTF-8 encoded.") from exc


def _single_page(text: str) -> ExtractedPage:
    return _ensure_text([ExtractedPage(page_number=1, text=text)])[0]


def _ensure_text(pages: list[ExtractedPage]) -> list[ExtractedPage]:
    if not any(page.text.strip() for page in pages):
        raise TextExtractionError("No extractable text found in document.")
    return pages
